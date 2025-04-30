# excel_manager.py

import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

def select_output_file_path():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Save the output Excel file"
    )
    return file_path

# Other functions remain unchanged...

def write_studies_to_excel(output_file_path, company, studies_data):
    print(f"Writing data for {company} to {output_file_path}")  # Debugging information
    if not studies_data:
        print(f"No data to write for {company}.")
        return

    if os.path.exists(output_file_path):
        workbook = load_workbook(output_file_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # If workbook is new, remove the default sheet

    sheet_title = company[:31]  # Excel sheet names have a maximum length of 31 characters
    if sheet_title in workbook.sheetnames:
        sheet = workbook[sheet_title]
    else:
        sheet = workbook.create_sheet(title=sheet_title)

    headers = list(studies_data[0].keys()) if studies_data else []

    # Write headers if the sheet is new or empty
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.append(headers)

    for study in studies_data:
        print(f"Study data: {study}")  # Detailed debugging information
        row = [','.join(item) if isinstance(item, list) else item for item in [study.get(header, '') for header in headers]]
        print(f"Appending row: {row}")  # Debugging information
        sheet.append(row)

    adjust_column_widths(sheet)
    workbook.save(output_file_path)
    print(f"Data for {company} successfully written.")

def adjust_column_widths(ws):
    """
    Adjusts the widths of columns based on their content.

    :param ws: The worksheet object.
    """
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width



